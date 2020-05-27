#!/usr/bin/env python
""" Estimate relative influence of tools and standards in Ten best practices for making reproducible biochemical models

:Author: Arthur Goldberg <Arthur.Goldberg@mssm.edu>
:Date: 2020-04-21
:Copyright: 2020, Karr Lab
:License: MIT
"""
from openpyxl import load_workbook
from pprint import pprint
from serpapi.google_scholar_search_results import GoogleScholarSearchResults
import ast
import bibtexparser
import collections
import csv
import datetime
import keys
import re
import requests
import subprocess
import sys
import time
import urllib.parse
import xml.etree.ElementTree as ET


# bibliography for the curated tools and standards
BIBLIOGRAPHY = 'guidelines_4_repro_models_2020__curated_standards.bib'
CURATED_STANDARDS_FILE = 'curated_standards.xlsx'
SURVEY_RESPONSES_FILE = 'paper_2018_curr_opin_sys_biol/survey_responses-edited.xlsx'
LATEX_TABLE_FILE = 'evaluated_standards.tex'
LATEX_TABLE_CITATIONS_FILE = 'evaluated_standards_citations.tex'
EVALUATED_STANDARDS_FILE = 'evaluated_standards.tsv'
TOOL_EMAIL = 'tool=mssm_citation_research&email=Arthur.Goldberg%40mssm.edu'
EUTILS = 'eutils.ncbi.nlm.nih.gov/entrez/eutils'
NCBI_ELINK_TEMPLATE = (f"https://{EUTILS}/elink.fcgi?dbfrom=pubmed&linkname=pubmed_pmc_ref"
                       f"s&id={{}}&{TOOL_EMAIL}")
NCBI_ESEARCH_TEMPLATE = f"https://{EUTILS}/esearch.fcgi?db=pubmed&term={{}}"
ESUMMARY_TEMPLATE = f"https://{EUTILS}/esummary.fcgi?db=pubmed&id={{}}&retmode=json&{TOOL_EMAIL}"


class GoogleScholar(object):

    def __init__(self):
        # To use SerpApi, create an account, get your private API key, create a keys.py file on the Python PATH, assign
        #   SERP_API_KEY = 'your private API key'
        # in keys.py. Keep keys.py secure.
        self.SERP_API_KEY = keys.SERP_API_KEY

    def get_gs_results(self, title, mock=False):
        """ Get num citations, publication year, and title from GS

        Args:
            title (:obj:`str`): paper title
            mock (:obj:`bool`): whether to return mock results

        Returns:
            :obj:`tuple`: Google Scholar title, number citations, publication year, errors
        """
        if mock:
            # test version of get_gs_results, which avoids using billable searches
            return '', len(title), 2000 + len(title)/10, []

        errors = []
        pub_year = None
        client = GoogleScholarSearchResults({"q": title, "serp_api_key": self.SERP_API_KEY})
        data = client.get_json()
        summary = data['organic_results'][0]['publication_info']['summary']
        PUB_YEAR_RE = '(\d\d\d\d)'
        match = re.search(PUB_YEAR_RE, summary)
        if match is None:
            errors.append(f"cannot get year for {title}")
        else:
            pub_year = int(match.group(1))
        num_citations = 0
        if "cited_by" in data['organic_results'][0]["inline_links"]:
            num_citations = data['organic_results'][0]["inline_links"]["cited_by"]["total"]
        gs_title = data['organic_results'][0]["title"]
        return gs_title, num_citations, pub_year, errors


class NCBIUtils(object):

    SLEEP_TIME = 2.0
    NCBI_API_KEY = None

    def __init__(self):
        # see https://ncbiinsights.ncbi.nlm.nih.gov/2017/11/02/new-api-keys-for-the-e-utilities/
        if hasattr(keys, 'NCBI_API_KEY'):
            self.NCBI_API_KEY = keys.NCBI_API_KEY
            self.SLEEP_TIME = 0.15

    def add_key(self, url):
        """ Add NCBI_API_KEY key to url, if it's known

        Args:
            url (:obj:`str`): url

        Returns:
            :obj:`str`: a URL, with NCBI_API_KEY query string entry, if it's known
        """
        if self.NCBI_API_KEY:
            return url + f'&api_key={self.NCBI_API_KEY}'
        return url

    def get_pm_id(self, title):
        """ Get PubMed ID for a title from the PutMed service, if possible

        Args:
            title (:obj:`str`): paper title

        Returns:
            :obj:`tuple`: the PubMed ID for title, errors if any
        """
        url = NCBI_ESEARCH_TEMPLATE.format(urllib.parse.quote(title))
        url = self.add_key(url)
        errors = []
        time.sleep(self.SLEEP_TIME)
        response = requests.get(url)
        if response.status_code != requests.codes.ok:
            errors.append(f"NCBI_ESEARCH_TEMPLATE request error {response.status_code} for title '{title}'")
            return None, errors
        root = ET.fromstring(response.text)
        pm_ids = [int(id.text) for id in root.iter('Id')]
        if len(pm_ids) == 0:
            return None, None
        if len(pm_ids) == 1:
            return pm_ids[0], None
        elif 1 < len(pm_ids):
            for pm_id in pm_ids:
                _, pm_title, errors = self.get_pub_metadata(pm_id)
                if errors:
                    return None, errors
                # ignore case in match
                if pm_title.casefold() == title.casefold():
                    return pm_id, None
            return None, None

    def get_num_citations(self, pm_id):
        """ Get PubMed citations for PubMed ID

        Args:
            pm_id (:obj:`str`): PubMed ID

        Returns:
            :obj:`tuple`: PubMed citations, errors if any
        """
        url = NCBI_ELINK_TEMPLATE.format(pm_id)
        url = self.add_key(url)
        errors = []
        time.sleep(self.SLEEP_TIME)
        response = requests.get(url)
        if response.status_code != requests.codes.ok:
            errors.append(f"request error {response.status_code} for pmc_id {pmc_id}")
            return None, errors
        return self.get_num_citations_from_xml(response.text), None

    @staticmethod
    def get_num_citations_from_xml(xml_string):
        """ Get number PubMed citations

        Args:
            xml_string (:obj:`str`): xml_string

        Returns:
            :obj:`int`: num PubMed citations
        """
        root = ET.fromstring(xml_string)
        num_citations = len(list(root.iter('Id'))) - 1
        return num_citations

    def get_pub_metadata(self, pm_id):
        """ Get PubMed metadata

        Args:
            pm_id (:obj:`str`): PubMed ID

        Returns:
            :obj:`tuple`: publication year, title, errors if any
        """
        url = ESUMMARY_TEMPLATE.format(pm_id)
        url = self.add_key(url)
        errors = []
        time.sleep(self.SLEEP_TIME)
        response = requests.get(url)
        if response.status_code != requests.codes.ok:
            errors.append(f"request error {response.status_code} for pm_id {pm_id}")
            return (None, errors)
        resp_dict = dict(ast.literal_eval(response.text))
        # get publication year
        pubdate = resp_dict["result"][str(pm_id)]["pubdate"]
        PUB_YEAR_RE = "(\d+)"
        match = re.search(PUB_YEAR_RE, pubdate)
        if match is None:
            errors.append(f"year not found in '{pubdate}' for {pm_id}")
        else:
            pub_year = int(match.group(1))
        # get title
        title = resp_dict["result"][str(pm_id)]["title"][:-1]
        return pub_year, title, errors


class Biblio(object):
    def __init__(self, filename):
        self.filename = filename
        with open(filename) as bibtex_file:
            self.bib_database = bibtexparser.load(bibtex_file)

    def get_entry_key(self, title):
        """ Get the bibliography key for a title

        Args:
            title (:obj:`str`): title

        Returns:
            :obj:`str`: the bibliography key
        """
        for citation in self.bib_database.entries:
            if citation['title'] == title:
                return citation['ID']


class CuratedStandards(object):

    # curated standard keys
    STANDARD = 'Standard / tool'
    TYPE = 'Type'
    TITLE = 'Title'
    BIB_KEY = 'bib_key'
    PUB_YEAR = 'pub_year'
    PM_ID = 'PM_id'
    PM_CITATIONS = 'PM_citations'
    GS_CITATIONS = 'GS_citations'
    SURVEY_ADOPTION_RATE = 'survey_adoption_rate'

    COLUMNS = (STANDARD,
               'Type of standard / tool',
               'Most cited paper',
               'Paper year',
               'PubMed (cites / yr)',
               'Scholar (cites / yr)',
               r'Reported use (\%)')

    def __init__(self, filename, biblio):
        self.filename = filename
        self.biblio = biblio
        self.curated_standards = self.spreadsheet_into_dicts(filename)

    @staticmethod
    def spreadsheet_into_dicts(filename):
        """ Read first worksheet in an Excel workbook with column headers into a list of dictionaries

        Args:
            filename (:obj:`str`): filename

        Returns:
            :obj:`list` of :obj:`dict`: data, in a list of rows, with a header-keyed dict for each row
        """
        workbook = load_workbook(filename, data_only=True)
        worksheet = workbook.active
        field_names = [col_name_cell.value for col_name_cell in worksheet[1]]
        records = []
        for row in worksheet.iter_rows(min_row=2):
            new_record = {}
            for field_name, cell in zip(field_names, row):
                new_record[field_name] = cell.value
            records.append(new_record)
        return records

    def check_all_titles(self):
        """ Ensure that all titles are in the bibliography
        """
        titles = self.read_curated_standards_column(self.TITLE)
        missing = set()
        for title in titles:
            if not self.biblio.get_entry_key(title):
                missing.add(title)
        if missing:
            print('Missing titles:', missing, file=sys.stderr)
        else:
            print(f"All {len(titles)} title(s) found in '{self.biblio.filename}'.")

    def read_curated_standards_column(self, col_name):
        """ Get all curated standards data in the given column

        Args:
            col_name (:obj:`str`): col_name

        Returns:
            :obj:`list`: all curated standards data in the given column
        """
        entries = []
        for curated_standard in self.curated_standards:
            if col_name in curated_standard:
                entries.append(curated_standard[col_name])
        return entries

    def enrich_with_pm_ids(self):
        """ Enrich the curated standards with PubMed IDs
        """
        missing_ids = []
        for curated_standard in self.curated_standards:
            title = curated_standard[self.TITLE]
            pm_id, errors = NCBIUtils().get_pm_id(curated_standard[self.TITLE])
            if pm_id is not None:
                curated_standard[self.PM_ID] = pm_id
            elif errors is None:
                missing_ids.append((title, 'no error'))
            else:
                missing_ids.append((title, errors))
        if missing_ids:
            print('Titles missing PM citations:', file=sys.stderr)
            pprint(missing_ids, stream=sys.stderr)
        else:
            print('All references have PM ids.')

    def enrich_with_num_pm_citations(self):
        """ Enrich the curated standards with the number of PubMed citations
        """
        missing_pm_cites = []
        for curated_standard in self.curated_standards:
            if self.PM_ID in curated_standard:
                pm_id = curated_standard[self.PM_ID]
                num_citations, errors = NCBIUtils().get_num_citations(pm_id)
                if num_citations is not None:
                    curated_standard[self.PM_CITATIONS] = num_citations
                else:
                    title = curated_standard[self.TITLE]
                    missing_pm_cites.append(title)
        if missing_pm_cites:
            print('Titles missing PM citations:', file=sys.stderr)
            pprint(missing_pm_cites, file=sys.stderr)
        else:
            print('All references with PM ids have PM citations.')

    def enrich_with_bib_key(self):
        """ Enrich the curated standards with bibliographic keys from the .bib file
        """
        missing_bib_keys = []
        for curated_standard in self.curated_standards:
            title = curated_standard[self.TITLE]
            entry_key = self.biblio.get_entry_key(title)
            if entry_key is not None:
                curated_standard[self.BIB_KEY] = entry_key
            else:
                missing_bib_keys.append(title)
        if missing_bib_keys:
            print('Titles missing bib key:', file=sys.stderr)
            pprint(missing_bib_keys, stream=sys.stderr)
        else:
            print('All references have bibliography keys.')

    def enrich_with_gs_data(self):
        """ Enrich the curated standards with number of citations and publication year from Google Scholar
        """
        missing_gs_data = []
        for curated_standard in self.curated_standards:
            title = curated_standard[self.TITLE]
            gs_title, num_citations, pub_date, errors = GoogleScholar().get_gs_results(title)
            if num_citations is None or pub_date is None or errors:
                missing_gs_data.append((title, errors))
            else:
                curated_standard[self.GS_CITATIONS] = num_citations
                curated_standard[self.PUB_YEAR] = pub_date
        if missing_gs_data:
            print('Titles not found on Google Scholar:', file=sys.stderr)
            pprint(missing_gs_data, stream=sys.stderr)
        else:
            print('All references found on Google Scholar.')

    def enrich_with_survey_data(self):
        """ Enrich the curated standards with survey data
        """
        survey = collections.defaultdict(list)
        for response in self.spreadsheet_into_dicts(SURVEY_RESPONSES_FILE):
            for col_name, value in response.items():
                survey[col_name].append(value)
        questions = \
            ['If you use models in your research, which tools do you most frequently use to build and/or simulate models?',
             'If you use models in your research, which resources do you most frequently use to distribute models?',
             'If you use models in your research, which languages do you most frequently use to represent models?']
        adoption = {}
        # number of subjects answering each question
        num_answers = collections.defaultdict(int)
        for question in questions:
            adoption[question] = collections.defaultdict(int)
            for responses in survey[question]:
                if responses:
                    num_answers[question] += 1
                    for tool in responses.split(';'):
                        # adoption is the number of uses of each tool, by question
                        adoption[question][tool] += 1
        fractional_adoption = {}
        for question in adoption:
            for tool, count in adoption[question].items():
                # some tools are responses to multiple questions; take the largest fractional adoption
                fractional_adoption[tool] = max(count/num_answers[question],
                                                fractional_adoption.get(tool, 0))
        for curated_standard in self.curated_standards:
            if curated_standard[self.STANDARD] in fractional_adoption:
                curated_standard[self.SURVEY_ADOPTION_RATE] = fractional_adoption[curated_standard[self.STANDARD]]
        print('Survey data incorporated.')

    @staticmethod
    def year_fraction(date):
        """ Convert date into fractional year

        From https://stackoverflow.com/a/36949905

        Args:
            date (:obj:`datetime`): date

        Returns:
            :obj:`float`: date as number of years including a fractional portion
        """
        start = datetime.date(date.year, 1, 1).toordinal()
        year_length = datetime.date(date.year+1, 1, 1).toordinal() - start
        return date.year + float(date.toordinal() - start) / year_length

    def write_evaluated_standards_file(self):
        """ Write the enriched, curated standards as a tsv file
        """
        evaluated_standards = self.generate_data_table()
        with open(EVALUATED_STANDARDS_FILE, 'w', newline='') as csvfile:
            evaluated_standards_writer = csv.writer(csvfile, delimiter='\t')
            evaluated_standards_writer.writerow(self.COLUMNS)
            for row in evaluated_standards:
                evaluated_standards_writer.writerow(row)

    def generate_data_table(self):
        """ Convert the enriched, curated standards into a list of rows

        Returns:
            :obj:`list` of :obj:`list`: the curated standards data, in a list of rows
        """
        n_columns = len(self.COLUMNS)
        current_year = self.year_fraction(datetime.datetime.today())
        rows = []
        for curated_standard in self.curated_standards:

            row = [''] * n_columns
            row[0] = curated_standard[self.STANDARD]
            row[1] = curated_standard[self.TYPE]
            row[2] = fr"\cite{{{curated_standard[self.BIB_KEY]}}}"

            if self.PUB_YEAR in curated_standard:
                row[3] = str(curated_standard[self.PUB_YEAR])
                age = current_year - curated_standard[self.PUB_YEAR]
                if age <= 0:
                    print(f"age ({age}) of '{curated_standard[self.TITLE]}' <= 0", file=sys.stderr)
                    continue

                # compute PM citations / year
                if self.PM_CITATIONS in curated_standard:

                    PM_cites_per_year = curated_standard[self.PM_CITATIONS] / age
                    row[4] = f"{PM_cites_per_year:.1f}"

                # compute GS citations / year
                if self.GS_CITATIONS in curated_standard:

                    GS_cites_per_year = curated_standard[self.GS_CITATIONS] / age
                    row[5] = f"{GS_cites_per_year:.1f}"

                if self.SURVEY_ADOPTION_RATE in curated_standard:
                    pct_adoption = curated_standard[self.SURVEY_ADOPTION_RATE] * 100.
                    row[6] = f"{pct_adoption:.1f}"

            rows.append(row)

        # sort rows by decreasing Google Scholar citations
        rows.sort(key=lambda curated_std: float(curated_std[5]), reverse=True)
        return rows

    LATEX_PACKAGES_AND_COMMANDS = r"""
% packages & commands used by "Standards and tools ordered by estimated influence" table
\usepackage{booktabs}
\usepackage{array}
\usepackage{longtable}
\usepackage{import}
% see: https://tex.stackexchange.com/a/119561
\newcolumntype{R}[1]{>{\raggedleft\arraybackslash}p{#1}}
\newcolumntype{L}[1]{>{\raggedright\arraybackslash}p{#1}}
"""
    COLUMN_ALIGNMENTS = ('L{2.2cm}',
                         'L{4cm}',
                         'L{1cm}',
                         'L{0.8cm}',
                         'R{1.1cm}',
                         'R{1cm}',
                         'R{1cm}')
    SMALL_COLUMNS = (True,
                     True,
                     False,
                     True,
                     True,
                     True,
                     True)

    def generate_latex_table(self, columns=None, column_alignments=None, small_columns=None):
        """ Convert the enriched, curated standards into a LaTeX table

        Args:
            columns (:obj:`tuple`): sequence of names of column headers
            column_alignments (:obj:`tuple`): sequence of column widths and alignments
            small_columns (:obj:`tuple` of :obj:`bool`): sequence indicating whether column text should be small

        Returns:
            :obj:`str`: the LaTeX table
        """
        if columns is None:
            columns=self.COLUMNS
        n_columns = len(columns)
        if column_alignments is None:
            column_alignments=self.COLUMN_ALIGNMENTS
        if small_columns is None:
            small_columns=self.SMALL_COLUMNS
        rows = self.generate_data_table()

        # select the data requested in columns in rows
        if columns != self.COLUMNS:
            chosen_cols = []
            for col in columns:
                chosen_cols.append(self.COLUMNS.index(col))
            filtered_rows = []
            for row in rows:
                new_row = []
                for cc in chosen_cols:
                    new_row.append(row[cc])
                filtered_rows.append(new_row)
            rows = filtered_rows

        tmp_rows = []
        for row in rows:
            sized_row = []
            for shrink, entry in zip(small_columns, row):
                if shrink:
                    entry = fr"\small{{{entry}}}"
                sized_row.append(entry)
            tmp_rows.append(sized_row)
        rows = tmp_rows

        CAPTION_INSTRUCTIONS = "%% Obtain caption from the file 'evaluated_standards_caption.txt'"
        CAPTION = '--Caption goes here.--'

        END_OF_LINE = r'\\' + '\n'
        HLINE = r'\hline' + '\n'
        TOPRULE = r'\toprule' + '\n'
        MIDRULE = r'\midrule' + '\n'

        TABLE_START = '\n' + r'\begin{longtable}'
        TABLE_END = r'\bottomrule\end{longtable}' + '\n'
        table = [TABLE_START]
        table.append('{' + ''.join(column_alignments) + '}\n')
        table.append(CAPTION_INSTRUCTIONS + '\n')
        table.append(fr"\caption{{{CAPTION}}}\\" + '\n')
        small_columns = [fr"\scriptsize{{{col}}}" for col in columns]

        header = []
        header.append(TOPRULE)
        header.append(' &'.join([fr"\textbf{{{col}}}" for col in small_columns]))
        header.append(r'\\' + '\n')

        table.append('% header for first page\n')
        table.extend(header)
        table.append(r'\endfirsthead' + '\n')

        table.append('% same header for subsequent pages\n')
        table.extend(header)
        table.append(MIDRULE)
        table.append(r'\endhead' + '\n')

        for row in rows:
            table.append(MIDRULE)
            table.append(' &'.join(row))
            table.append(END_OF_LINE)

        table.append(TABLE_END)
        complete_table = ''.join(table)
        return complete_table

    def output_latex_table(self, filename=LATEX_TABLE_FILE):
        with open(filename, 'w') as latex_table:
            latex_table.write(self.generate_latex_table())

    def output_latex_table_of_citations(self, filename=LATEX_TABLE_CITATIONS_FILE):
        with open(filename, 'w') as latex_table:
            latex_table.write(self.generate_latex_table(columns=['Most cited paper'],
                                                        column_alignments=['L{16cm}'],
                                                        small_columns=[False]))


def main():
    """ Create the evaluated standards files
    """
    biblio = Biblio(BIBLIOGRAPHY)
    curated_standards = CuratedStandards(CURATED_STANDARDS_FILE, biblio)
    curated_standards.check_all_titles()
    curated_standards.enrich_with_bib_key()
    curated_standards.enrich_with_survey_data()
    curated_standards.enrich_with_gs_data()
    curated_standards.enrich_with_pm_ids()
    curated_standards.enrich_with_num_pm_citations()
    curated_standards.write_evaluated_standards_file()
    curated_standards.output_latex_table()
    curated_standards.output_latex_table_of_citations()

if __name__ == '__main__':
    main()
